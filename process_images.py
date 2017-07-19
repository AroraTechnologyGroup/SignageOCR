import os
from PIL import Image
import pytesseract
from openpyxl import load_workbook
input_excel = r"Y:\Projects\DFW\DFW Web GDB\Text_OCR.xlsx"
wb = load_workbook(input_excel)
ws = wb.active

pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files (x86)\\Tesseract-OCR\\tesseract'
input_dir = r"Y:\Projects\DFW\DFW Web GDB\INTERIOR_Photos"

files = os.listdir(input_dir)
for file in files:
    try:
        # locate the row in the excel table using the filename and update the text column
        text = pytesseract.image_to_string(Image.open(os.path.join(input_dir, file)))
        text = text.strip()
        text = text.replace('\n', ',')
        splits = text.split(",")
        new_text = []
        for lets in splits:
            new_let = lets.strip()
            if new_let:
                new_text.append(new_let)
        if new_text:
            text = ", ".join(new_text)
            for row in ws.iter_rows(min_row=2, max_col=4):
                if row[0].value == file:
                    row[3].value = text
                    print(file, text)
    except Exception as e:
        print(e)
wb.save(input_excel)
